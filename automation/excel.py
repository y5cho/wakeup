from openpyxl import Workbook 
from random import *
from openpyxl.utils.cell import coordinate_from_string
#-----------------prep-------------
wb = Workbook() 
ws = wb.active 
ws.title = "sheet1"
#-----------------sheet-------------
second_worksheet = wb.create_sheet("sheet2")
third_worksheet = wb.create_sheet("sheet3")

second_worksheet.sheet_properties.tabColor = "ff66ff" #RGB color 
#------------------------------
ws2 = wb["sheet2"]   #necessary since sheet1 is not the first sheet we created by making it active 
#------------------------------
ws.append(["number", "english", "math"])
for i in range(1,11):
    ws.append([i, randint(0, 100), randint(0,100)])

'''
col_B = ws["B"]
for cell in col_B:
    print(cell.value)

col_range = ws["B:C"]
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1]
for cell in row_title:
    print(cell.value)

row_range = ws[2:ws.max_row]
for rows in row_range:
    for row in rows:
        print(row.coordinate, end=" ")
        print(row.value, end=" ")
    print()
'''
#all rows 
for row in tuple(ws.rows):
    print(row[2].value)

for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[0].value, row[1].value)


#all columns 
tuple(ws.columns)
#------------------------------
ws2["A1"] = "Test"
print(ws2["A1"].value)  #= ws2.cell(row=1, column=1).value
index = 1
for x in range(1,11):
    for y in range(1,11):
        ws2.cell(row=x, column=y, value=index)
        index = index + 1

#-----------------sheet-------------
wb.save("sample.xlsx")
wb.close()


#target = wb.copy_worksheet(ws2)
#target.title = "copied sheet"