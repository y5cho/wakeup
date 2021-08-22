from openpyxl import load_workbook
from random import *

from openpyxl.workbook.workbook import Workbook

wb = load_workbook("Book1.xlsx")
sheets = wb.sheetnames
Sheet1 = wb[sheets[0]]
Sheet2 = wb.create_sheet("Sheet2")




for y in range(2,12):
    Sheet1.cell(row = y, column = 2).value = randint(1,100)
    Sheet1.cell(row = y, column = 3).value = randint(1,100)


wb.save("Book1.xlsx")