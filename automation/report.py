from openpyxl import load_workbook
'''
from random import *
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils.cell import coordinate_from_string
'''
file = load_workbook("Book1.xlsx")
sheets = file.sheetnames 
Sheet1 = file[sheets[0]]
col_B = Sheet1["B"]

for row in Sheet1.iter_rows(min_row = 1, max_row=11, min_col=1, max_col=3):
    print(row[0].value, row[1].value, row[2].value)

